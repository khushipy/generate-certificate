import os
import sys
import openpyxl
import subprocess
from concurrent.futures import ProcessPoolExecutor, wait, FIRST_COMPLETED
from datetime import datetime
import multiprocessing
import traceback

EXCEL_FILENAME = "input_file.xlsx"
CONFIG_FILENAME = "input.txt"  # First line: number of input columns; Second line: executable path/name
STATUS_COL_HEADERS = ["Status", "Start Time", "End Time", "CPU Core Used", "Output"]


def find_available_cores():
    total = multiprocessing.cpu_count()
    return max(1, total - 2)


def load_config(path):
    try:
        with open(path, "r") as f:
            lines = [line.strip() for line in f if line.strip()]
        if len(lines) < 2:
            raise ValueError(f"The config file '{path}' must have at least two lines: input column count and exe path")
        num_columns = int(lines[0])
        exe_name = lines[1]
        return num_columns, exe_name
    except Exception as e:
        print(f"[ERROR] Unable to read config from {path}: {e}")
        sys.exit(1)


def ensure_status_columns(ws, start_col):
    # Create the fixed status columns starting at start_col
    for i, col_name in enumerate(STATUS_COL_HEADERS):
        ws.cell(row=1, column=start_col + i, value=col_name)
    return {
        "status": start_col,
        "start_time": start_col + 1,
        "end_time": start_col + 2,
        "core": start_col + 3,
        "output": start_col + 4,
        "next_after_status": start_col + len(STATUS_COL_HEADERS)
    }


def reset_running_on_resume(ws, status_col):
    for row in ws.iter_rows(min_row=2, min_col=status_col, max_col=status_col):
        if str(row[0].value).strip().lower() == "running":
            row[0].value = "pending"


def run_exe_on_batch(inputs, exe_name):
    try:
        exe_full_path = os.path.abspath(exe_name)
        args = [exe_full_path] + [str(x) for x in inputs]
        print(f"[DEBUG] Running: {' '.join(args)}")
        result = subprocess.run(args, capture_output=True, text=True, check=True)
        return result.stdout, result.stderr
    except Exception as e:
        tb = traceback.format_exc()
        stderr = getattr(e, "stderr", "") or ""
        print(f"[ERROR] Exception during exe run: {e}\n{tb}")
        return f"[ERROR] Exception: {e}\n{tb}\n{stderr}", stderr


def process_batch(row_idx, batch_inputs, exe_name, core_no):
    start = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    outs, stderr = run_exe_on_batch(batch_inputs, exe_name)
    end = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    case_id = batch_inputs[11] if len(batch_inputs) > 11 and batch_inputs[11] else f"UNKNOWN_{row_idx}"
    out_fname = f"{case_id}.txt"
    try:
        with open(out_fname, "w", encoding="utf-8") as f:
            f.write(outs)
            if stderr:
                f.write(f"\n[STDERR]:\n{stderr}")
    except Exception as e:
        outs += f"\n[ERROR writing output file]: {e}"

    return (row_idx, "completed", start, end, f"Core {core_no + 1}", outs)


def main():
    base_dir = os.path.abspath(os.path.dirname(__file__))
    excel_path = os.path.join(base_dir, EXCEL_FILENAME)
    config_path = os.path.join(base_dir, CONFIG_FILENAME)

    if not os.path.exists(excel_path):
        print(f"[ERROR] Excel file not found: {excel_path}")
        sys.exit(1)

    if not os.path.exists(config_path):
        print(f"[ERROR] Config file not found: {config_path}")
        sys.exit(1)

    input_col_count, exe_name = load_config(config_path)
    n_cores = find_available_cores()

    print(f"[INFO] Using {n_cores} CPU cores (2 kept idle).")
    print(f"[INFO] Processing first {input_col_count} columns per batch.")
    print(f"[INFO] Executable: {exe_name}")

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Status columns start immediately after input columns
    status_cols = ensure_status_columns(ws, input_col_count + 1)
    status_col = status_cols["status"]
    output_start_col = status_cols["next_after_status"]

    reset_running_on_resume(ws, status_col)
    wb.save(excel_path)

    rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False))
    batches = []
    for i, row in enumerate(rows):
        batch_inputs = [row[j].value if j < input_col_count else "" for j in range(input_col_count)]
        excel_row_num = i + 2
        batches.append((excel_row_num, batch_inputs))

    # Mark empty or unknown statuses as pending
    for row_num, _ in batches:
        status_cell = ws.cell(row=row_num, column=status_col)
        val = str(status_cell.value).strip().lower() if status_cell.value else ""
        if val not in ("pending", "running", "completed"):
            status_cell.value = "pending"
    wb.save(excel_path)

    pending_batches = [(r, b) for (r, b) in batches if str(ws.cell(row=r, column=status_col).value).strip().lower() == "pending"]

    futures = {}
    submitted_indices = set()
    next_to_submit = 0

    with ProcessPoolExecutor(max_workers=n_cores) as executor:
        # Submit the initial batches up to n_cores
        while next_to_submit < min(n_cores, len(pending_batches)):
            row_idx, batch_inputs = pending_batches[next_to_submit]
            ws.cell(row=row_idx, column=status_col).value = "running"
            wb.save(excel_path)
            fut = executor.submit(process_batch, row_idx, batch_inputs, exe_name, next_to_submit % n_cores)
            futures[fut] = row_idx
            submitted_indices.add(next_to_submit)
            next_to_submit += 1

        while futures:
            done, _ = wait(futures.keys(), return_when=FIRST_COMPLETED)
            for fut in done:
                row_idx = futures[fut]
                try:
                    r_idx, status, start, end, core_str, full_output = fut.result()
                except Exception as e:
                    print(f"[ERROR] Exception in batch at row {row_idx}: {e}")
                    ws.cell(row=row_idx, column=status_col).value = "error"
                    wb.save(excel_path)
                    futures.pop(fut)
                    continue

                # Update status columns
                ws.cell(row=r_idx, column=status_col).value = status
                ws.cell(row=r_idx, column=status_col + 1).value = start
                ws.cell(row=r_idx, column=status_col + 2).value = end
                ws.cell(row=r_idx, column=status_col + 3).value = core_str

                # Clear "Output" status column (do not put whole output string here)
                ws.cell(row=r_idx, column=status_col + 4).value = ""

                # Split first line of output into tab-separated fields, write after Status columns
                first_line = full_output.strip().split('\n')[0] if full_output else ""
                output_fields = first_line.split('\t') if first_line else []

                # Clear previous output fields columns (optional)
                max_output_clear = 50
                for offset in range(max_output_clear):
                    ws.cell(row=r_idx, column=output_start_col + offset).value = None

                for i, val in enumerate(output_fields):
                    ws.cell(row=r_idx, column=output_start_col + i).value = val

                wb.save(excel_path)

                futures.pop(fut)

                # Submit the next batch if available
                if next_to_submit < len(pending_batches):
                    next_row_idx, next_batch_inputs = pending_batches[next_to_submit]
                    ws.cell(row=next_row_idx, column=status_col).value = "running"
                    wb.save(excel_path)
                    f = executor.submit(process_batch, next_row_idx, next_batch_inputs, exe_name, next_to_submit % n_cores)
                    futures[f] = next_row_idx
                    submitted_indices.add(next_to_submit)
                    next_to_submit += 1

    wb.save(excel_path)
    print("[INFO] All batches processed successfully.")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"[FATAL ERROR]: {e}")
        traceback.print_exc()
